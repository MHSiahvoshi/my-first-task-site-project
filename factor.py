try:
    from openpyxl import *
    import my_sql
    import datetime
    import pandas as pd
except Exception as x:
    print(x)

wb = Workbook()
wb.save('temporary factor.xlsx')


def create_serial_factor(code: str):
    # now = datetime.datetime.now().time()
    time_today = datetime.datetime.today()
    out = code + '_' + str(time_today)
    return out


def get_max_row(sheet):
    # count = 0
    # for row in sheet:
    #     if not all([cell is None for cell in row]):
    #         count += 1
    out = sheet.max_row
    return out


def get_total_products(sheet):
    column_index = 5
    column_values = [cell.value for col in sheet.iter_cols(min_col=column_index, max_col=column_index) for cell in col
                     if cell.value is not None]
    # print(column_values)
    out = sum(column_values[1::])

    # print(column_values)
    return out


def manage_time():
    out = []
    out_flag = False
    mtf = load_workbook('temporary factor.xlsx')
    for sheet in mtf.sheetnames:
        ws = mtf.get_sheet_by_name(sheet)
        if ws[f'B{get_max_row(ws)}'].value == datetime.datetime.today().day:
            if ws[f'G{get_max_row(ws)}'].value == 0:
                mtf.remove_sheet(ws)
            else:
                out.append((ws, ws[f'C{get_max_row(ws)}'], ws[f'E{get_max_row(ws)}']))
                mtf.remove_sheet(ws)
                out_flag = True

    if len(mtf.sheetnames) == 0:
        mtf.create_sheet('sheet')
        mtf.save('temporary factor.xlsx')
    else:
        mtf.save('temporary factor.xlsx')
    if out_flag:
        return out
    else:
        return False


def add_product_to_tf(data: list):
    ptf = load_workbook('temporary factor.xlsx')
    current = ptf.active
    data_for_xlsx = []
    data_for_xlsx.extend(list(my_sql.get_product_detail([data[1]])[0]))
    # print(current)
    if current.title == 'Sheet':
        # print('ok')
        current.title = f'customer_code = {data[0]}'
        current.delete_rows(get_max_row(current))
        current.append(('product_code', 'product_name', 'price', 'number', 'total'))
        created_date = datetime.date.today()
        expire_date = created_date.day + 0
        data_for_xlsx.append(data[2])
        data_for_xlsx.append(data_for_xlsx[2] * data_for_xlsx[3])
        current.append(tuple(data_for_xlsx))
        max_row = get_max_row(current)
        current[f'A{max_row + 1}'] = "expire data"
        current[f'B{max_row + 1}'] = expire_date
        current[f'C{max_row + 1}'] = create_serial_factor(f'{data[0]}{data[0]}')
        current[f'D{max_row + 1}'] = "final total"
        current[f'E{max_row + 1}'] = get_total_products(current)
        current[f'F{max_row + 1}'] = "is_paid"
        current[f'G{max_row + 1}'] = 0

    else:
        if f'customer_code = {data[0]}' not in ptf.sheetnames:
            ptf.create_sheet(f'customer_code = {data[0]}')
            # select_ws = wb.
            select_ws = ptf.get_sheet_by_name(f'customer_code = {data[0]}')
            select_ws.delete_rows(get_max_row(select_ws))
            select_ws.append(('product_code', 'product_name', 'price', 'number', 'total'))
            created_date = datetime.date.today()
            expire_date = created_date.day + 0
            data_for_xlsx.append(data[2])
            data_for_xlsx.append(data_for_xlsx[2] * data_for_xlsx[3])
            select_ws.append(tuple(data_for_xlsx))
            max_row = get_max_row(select_ws)
            select_ws[f'A{max_row + 1}'] = "expire data"
            select_ws[f'B{max_row + 1}'] = expire_date
            select_ws[f'C{max_row + 1}'] = create_serial_factor(f'{data[0]}{data[0]}')
            select_ws[f'D{max_row + 1}'] = "final total"
            select_ws[f'E{max_row + 1}'] = get_total_products(select_ws)
            select_ws[f'F{max_row + 1}'] = "is_paid"
            select_ws[f'G{max_row + 1}'] = 0
            # create_time(code=f'{data[0]}',
            #             serial=create_serial_factor(f'{data[0]}{data[0]}'), expire=expire_date)

        else:
            # current.append(('product_code', 'product_name', 'price', 'number', 'total'))
            select_ws = ptf.get_sheet_by_name(f'customer_code = {data[0]}')
            expire_date = select_ws[f'B{get_max_row(select_ws)}'].value
            select_ws.delete_rows(get_max_row(select_ws))
            data_for_xlsx.append(data[2])
            data_for_xlsx.append(data_for_xlsx[2] * data_for_xlsx[3])
            select_ws.append(tuple(data_for_xlsx))
            max_row = get_max_row(select_ws)
            select_ws[f'A{max_row + 1}'] = "expire data"
            select_ws[f'B{max_row + 1}'] = expire_date
            select_ws[f'C{max_row + 1}'] = create_serial_factor(f'{data[0]}{data[0]}')
            select_ws[f'D{max_row + 1}'] = "final total"
            select_ws[f'E{max_row + 1}'] = get_total_products(select_ws)
            select_ws[f'F{max_row + 1}'] = "is_paid"
            select_ws[f'G{max_row + 1}'] = 0

    ptf.save('temporary factor.xlsx')


# def log_factor():
#     if manage_time():
#         my_sql.register_factor(manage_time())


def get_basket(data):
    # df = pd.read_excel('temporary factor.xlsx', sheet_name=f'customer_code = {data}')
    gb = load_workbook('temporary factor.xlsx')
    select_ws = gb.get_sheet_by_name(f'customer_code = {data}')
    temp = []
    for row in select_ws.iter_rows(values_only=True):
        temp.append(row)
    df = pd.Series(temp[1: -1: 1])
    return df


def paid(*args):
    ptf = load_workbook('temporary factor.xlsx')
    select_ws = ptf.get_sheet_by_name(f'customer_code = {args[0]}')
    select_ws[f'G{get_max_row(select_ws)}'] = args[1]
    ptf.save('temporary factor.xlsx')
    ptf.close()


def delete_product(*args):
    dp = load_workbook('temporary factor.xlsx')
    select_ws = dp.get_sheet_by_name(f'customer_code = {args[0]}')
    select_ws.delete_rows(args[1] + 2)
    select_ws[f'E{get_max_row(select_ws)}'] = get_total_products(select_ws)
    if get_max_row(select_ws) == 2:
        dp.remove_sheet(select_ws)
        if len(dp.sheetnames) == 0:
            dp.create_sheet('sheet')
    dp.save('temporary factor.xlsx')
    dp.close()


def send_data_to_sql(data: list):

    my_sql.register_factor(data)


def check_if_paid():
    cip = load_workbook('temporary factor.xlsx')
    for sheet in cip.sheetnames:
        select_ws = cip.get_sheet_by_name(sheet)
        if select_ws[f'G{get_max_row(select_ws)}'].value == 1:
            print(int(select_ws.title[-1]))
            print(select_ws[f'E{get_max_row(select_ws)}'].value)
            print(str(datetime.date.today()))
            print(type(select_ws.title[-1]))
            temp = int(str(select_ws.title)[-1])
            print(temp, type(temp))
            send_data_to_sql([temp,
                              select_ws[f'E{get_max_row(select_ws)}'].value,
                              str(datetime.date.today())[0: 12]])
            cip.remove_sheet(select_ws)
            if len(cip.sheetnames) == 0:
                cip.create_sheet('Sheet')
        cip.save('temporary factor.xlsx')
    cip.close()

# todo: manage pay if paid remove factor
# todo: calculate final total if any change happened in factor
# todo: mp4 file fo fun:)
