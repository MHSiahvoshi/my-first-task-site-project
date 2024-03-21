import openpyxl
import datetime

card_num_check = False
card_cvv2_check = False
card_expiration_date_check = False
payment = False


class Payment:
    def card(self, value):
        global card_num_check, card_cvv2_check, payment
        sum = 0
        if value.isdigit() and len(value) == 16:
            for i in range(len(value)):
                if (i+1) % 2 == 0:
                    result = int(value[i]) * 1
                else:
                    if (int(value[i]) * 2) > 9:
                        result = (int(value[i]) * 2) - 9
                    else:
                        result = int(value[i]) * 2
                sum += result
            if sum != 0:
                if sum % 10 == 0:
                    print("شماره کارت صحیح میباشد")
                    card_num_check = True
                    x = value[0:6]

                    # for j in range(0, 7):
                    #     x.join(value[j])
                    # print(x)

                    path = "bank.xlsx"
                    wb_obj = openpyxl.load_workbook(path)
                    sheet_obj = wb_obj.active
                    rows = sheet_obj.max_row
                    # lst_reows = []
                    # for i in rows:
                    #     lst_reows.append(i)

                    # print (lst_reows)
                    for i in range(1, rows + 1):
                        cell_obj = sheet_obj.cell(row=i, column=2)
                        # print (cell_obj.value)
                        # print(x)
                        # print(len(x))
                        # print(len(cell_obj.value))
                        if x == str(cell_obj.value):
                            print(sheet_obj.cell(row=i, column=1).value)

                            return card_num_check
                        else:
                            print("این شماره کارت در ایران وجود ندارد")
                    else:
                        print("شماره کارت صحیح نمیباشد")
        else:
            print("لطفا شماره کارت صحیح وارر کنید")

    def cvv2(self, value: str):
        global card_cvv2_check
        if value.isdigit() and len(value) <= 4:
        # for i in range(4):
        #     if value[i].isdigit() and 0 < int(value[i]) <= 9:
        #     print('ok')
            card_cvv2_check = True
            # return card_cvv2_check
        else:
            print("cvv2 اشتباه است")

    def expiration_date(self, m, y):
        global payment, card_expiration_date_check
        date = datetime.datetime.now()
        if int(y) > date.year or (int(y) == date.year and int(m) > date.month):
            card_expiration_date_check = True
            payment = True
            return card_expiration_date_check
        else:
            print("تاریخ گذشته است")


def pay():
    card_num = input("لطفا شماره کارت را وارد کنید ")
    pay = Payment()
    pay.card(card_num)
    if card_num_check == True:
        ccv2 = input("لطفا cvv2 را وارد کنید")
        pay.cvv2(ccv2)
        if card_cvv2_check == True:
            daate = input("لطفا سال را وارد کنید")
            daate2 = input("لطفا ماه را وارد کنید")
            pay.expiration_date(daate2, daate)
            if card_expiration_date_check == True:
                return payment
            else:
                # payment = False
                return False
